import openpyxl
import json

# Đường dẫn đến file Excel
excel_file_path = "questions.xlsx"
json_file_path = "data.json"

# Mở file Excel
wb = openpyxl.load_workbook(excel_file_path)
sheet1 = wb["sheet1"]
sheet2 = wb["sheet2"]

# Tạo danh sách lưu dữ liệu
data = []

# Duyệt qua các dòng, nhóm theo cặp (dòng lẻ là câu hỏi, dòng chẵn là đáp án)
for i in range(1, sheet2.max_row + 1):
    question = sheet1.cell(row=i, column=1).value
    answer = sheet2.cell(row=i, column=1).value

    if question and answer:  # Chỉ thêm nếu cả câu hỏi và đáp án đều có giá trị
        data.append({
            "question": question[1:].strip() if len(question) > 1 else "",
            "answer": answer[1:].strip() if len(answer) > 1 else ""
        })

# Ghi dữ liệu ra file JSON
with open(json_file_path, "w", encoding="utf-8") as json_file:
    json.dump(data, json_file, ensure_ascii=False, indent=2)

print(f"Dữ liệu đã được chuyển đổi và lưu vào {json_file_path}")
